"""Geração do relatório Excel segregado pelas classificações RN01-RN12."""

from __future__ import annotations

from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.chart import DoughnutChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.excel_reporting.models import RegistroValidado
from src.ml_audit import MLDecisionAudit
from src.excel_reporting.validation_service import (
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO_ENTRADA,
    CLASSIFICACAO_VALIDO,
    MOTIVOS,
)
from src.operational_indicators import (
    OperationalIndicators,
    _percentual,
)

REPORT_SHEET_NAMES = (
    "Resumo",
    "Todos",
    "Válidos",
    "Divergências",
    "Ambíguos",
    "Erros de Entrada",
    "Ranking de Regras",
    "Dicionário",
    "Decisões de ML",
)

CLASSIFICATION_SHEETS = {
    CLASSIFICACAO_VALIDO: "Válidos",
    CLASSIFICACAO_DIVERGENCIA: "Divergências",
    CLASSIFICACAO_AMBIGUO: "Ambíguos",
    CLASSIFICACAO_ERRO_ENTRADA: "Erros de Entrada",
}

DATA_SHEET_NAMES = ("Todos", *CLASSIFICATION_SHEETS.values())
RANKING_SHEET_NAME = "Ranking de Regras"
DICTIONARY_SHEET_NAME = "Dicionário"
ML_DECISIONS_SHEET_NAME = "Decisões de ML"

BUSINESS_COLUMNS = (
    "Data de referência",
    "Lote",
    "Produto",
    "Linha",
    "Turno",
    "Status",
    "Responsável",
    "Observação",
    "Classificação",
    "Motivo",
)

RANKING_COLUMNS = (
    "Código da Regra",
    "Nome / Descrição da Regra",
    "Total de Ocorrências",
    "% do Total",
)

DICTIONARY_COLUMNS = (
    "Categoria",
    "Termo",
    "Definição",
    "Fórmula / Meta de Referência",
)

ML_DECISION_COLUMNS = (
    "Data/hora",
    "ID da execução",
    "ID do bot",
    "Lote",
    "Classe prevista",
    "Probabilidade",
    "Nível de confiança",
    "Ação",
    "Resultado aplicado",
    "Latência (ms)",
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUMMARY_TITLE_FILL = PatternFill(fill_type="solid", fgColor="17365D")
MAX_DASHBOARD_DAYS = 10

CLASSIFICATION_COLORS = {
    CLASSIFICACAO_VALIDO: "70AD47",
    CLASSIFICACAO_DIVERGENCIA: "C00000",
    CLASSIFICACAO_AMBIGUO: "F4B183",
    CLASSIFICACAO_ERRO_ENTRADA: "7F8C8D",
}

SUMMARY_CARDS = (
    ("A4:C4", "A5:C6", "Total de registros", "total_registros", "1F4E78", False),
    ("E4:G4", "E5:G6", "Total de válidos", "validos_qtd", "70AD47", False),
    ("I4:K4", "I5:K6", "% de válidos", "validos_pct", "A9D18E", True),
    (
        "A8:C8",
        "A9:C10",
        "Total de divergências",
        "divergencias_qtd",
        "C00000",
        False,
    ),
    (
        "E8:G8",
        "E9:G10",
        "% de divergências",
        "divergencias_pct",
        "E26B6B",
        True,
    ),
    ("I8:K8", "I9:K10", "Total de ambíguos", "ambiguos_qtd", "ED7D31", False),
    ("A12:C12", "A13:C14", "% de ambíguos", "ambiguos_pct", "F4B183", True),
    (
        "E12:G12",
        "E13:G14",
        "Total de erros de entrada",
        "erros_entrada_qtd",
        "5B6573",
        False,
    ),
    (
        "I12:K12",
        "I13:K14",
        "% de erros de entrada",
        "erros_entrada_pct",
        "A5A5A5",
        True,
    ),
)

EXECUTIVE_CARDS = (
    ("M4:P4", "M5:P6", "6. Regra mais acionada", "4472C4"),
    ("M8:P8", "M9:P10", "7. Taxa de qualidade da entrada · meta > 80%", "548235"),
    ("M12:P12", "M13:P14", "8. Taxa de revisão humana · meta < 15%", "BF9000"),
    ("M16:P16", "M17:P18", "9. Taxa de retrabalho · meta < 6%", "C65911"),
    ("M20:P20", "M21:P22", "10. Ganho estimado de tempo", "5B9BD5"),
)

RESERVED_RULE_DESCRIPTION = (
    "Regra preservada pelo contrato da atividade, sem ocorrência específica "
    "na massa atual"
)


def write_excel_report(
    ordered_records: list[RegistroValidado],
    indicators: OperationalIndicators,
    output_path: str | Path,
    ml_decisions: Iterable[MLDecisionAudit] = (),
) -> Path:
    """Grava o workbook executivo com nove abas e auditoria opcional de ML."""
    destination = Path(output_path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    _validate_classifications(ordered_records)
    all_rows = [_business_row(record) for record in ordered_records]
    ranking_rows = _ranking_rows(ordered_records)
    dictionary_rows = _dictionary_rows()
    ml_decision_rows = [_ml_decision_row(decision) for decision in ml_decisions]

    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="Resumo", index=False)
        _frame_from_rows(all_rows).to_excel(writer, sheet_name="Todos", index=False)

        for classification, sheet_name in CLASSIFICATION_SHEETS.items():
            rows = [
                row
                for record, row in zip(ordered_records, all_rows, strict=True)
                if record.classificacao == classification
            ]
            _frame_from_rows(rows).to_excel(writer, sheet_name=sheet_name, index=False)

        pd.DataFrame(ranking_rows, columns=RANKING_COLUMNS).to_excel(
            writer,
            sheet_name=RANKING_SHEET_NAME,
            index=False,
        )
        pd.DataFrame(dictionary_rows, columns=DICTIONARY_COLUMNS).to_excel(
            writer,
            sheet_name=DICTIONARY_SHEET_NAME,
            index=False,
        )
        pd.DataFrame(ml_decision_rows, columns=ML_DECISION_COLUMNS).to_excel(
            writer,
            sheet_name=ML_DECISIONS_SHEET_NAME,
            index=False,
        )

        workbook = writer.book
        _format_summary_sheet(workbook["Resumo"], ordered_records, indicators)
        for sheet_name in DATA_SHEET_NAMES:
            _format_data_sheet(workbook[sheet_name])
        _format_table_sheet(
            workbook[RANKING_SHEET_NAME],
            table_name="TabelaRankingRegras",
            percentage_columns=(4,),
        )
        _format_table_sheet(
            workbook[DICTIONARY_SHEET_NAME],
            table_name="TabelaDicionario",
        )
        _format_ml_decisions_sheet(workbook[ML_DECISIONS_SHEET_NAME])

    return destination


def _ml_decision_row(decision: MLDecisionAudit) -> dict[str, Any]:
    return dict(
        zip(
            ML_DECISION_COLUMNS,
            (
                decision.timestamp,
                decision.execution_id,
                decision.bot_id,
                decision.lote_id,
                decision.classe,
                decision.probabilidade,
                decision.nivel_confianca,
                decision.acao,
                decision.resultado_aplicado,
                decision.latencia_ms,
            ),
            strict=True,
        )
    )


def _format_ml_decisions_sheet(sheet: Any) -> None:
    _format_table_sheet(sheet, table_name="TabelaDecisoesML")
    for cell in sheet["F"][1:]:
        cell.number_format = "0.0000"
    for cell in sheet["J"][1:]:
        cell.number_format = "0.000"


def _ranking_rows(records: list[RegistroValidado]) -> list[dict[str, Any]]:
    counts = Counter(
        record.regra_aplicada for record in records if record.regra_aplicada
    )
    total = len(records)
    return [
        {
            RANKING_COLUMNS[0]: code,
            RANKING_COLUMNS[1]: MOTIVOS.get(code, "Regra desconhecida"),
            RANKING_COLUMNS[2]: count,
            RANKING_COLUMNS[3]: _percentual(count, total) / 100.0,
        }
        for code, count in counts.most_common()
    ]


def _dictionary_rows() -> list[dict[str, str]]:
    rows = [
        _dictionary_row(
            "Termo operacional",
            "Lote",
            "Conjunto de itens de produção identificado por um código único.",
            "Campo usado para conferência e consulta à base de referência.",
        ),
        _dictionary_row(
            "Termo operacional",
            "Base de Referência",
            "Cadastro controlado dos lotes reconhecidos pelo processo.",
            "Um lote ausente aciona a RN05.",
        ),
        _dictionary_row(
            "Termo operacional",
            "Status normalizado",
            "Forma padronizada do resultado informado na entrada.",
            "OK vira APROVADO e NOK vira REPROVADO.",
        ),
        _dictionary_row(
            "Termo operacional",
            "Regra aplicada",
            "Regra principal que determinou a classificação do registro.",
            "Respeita a precedência Erro de Entrada > Divergência > Ambíguo.",
        ),
        _dictionary_row(
            "Termo operacional",
            "% do Total",
            "Participação das ocorrências de uma regra no lote processado.",
            "Ocorrências da regra ÷ total de registros × 100.",
        ),
        _dictionary_row(
            "Classificação",
            CLASSIFICACAO_VALIDO,
            "Registro que não acionou nenhuma das regras RN01–RN12.",
            "Sem regra acionada.",
        ),
        _dictionary_row(
            "Classificação",
            CLASSIFICACAO_DIVERGENCIA,
            "Registro com conflito de referência, reprovação sem observação ou duplicidade diária.",
            "RN05, RN10 ou RN11.",
        ),
        _dictionary_row(
            "Classificação",
            CLASSIFICACAO_AMBIGUO,
            "Registro cujo status não permite decisão automática segura.",
            "RN09; encaminhar para revisão humana.",
        ),
        _dictionary_row(
            "Classificação",
            CLASSIFICACAO_ERRO_ENTRADA,
            "Registro incompleto ou com data inválida.",
            "RN01–RN04 ou RN12.",
        ),
        _dictionary_row(
            "Indicador",
            "Total de registros",
            "Quantidade total de lotes processados.",
            "Contagem de todos os registros.",
        ),
        _dictionary_row(
            "Indicador",
            "Válidos",
            "Quantidade e percentual de registros válidos.",
            "Válidos ÷ total × 100.",
        ),
        _dictionary_row(
            "Indicador",
            "Divergências",
            "Quantidade e percentual de registros divergentes.",
            "Divergências ÷ total × 100.",
        ),
        _dictionary_row(
            "Indicador",
            "Ambíguos",
            "Quantidade e percentual de registros que exigem decisão humana.",
            "Ambíguos ÷ total × 100.",
        ),
        _dictionary_row(
            "Indicador",
            "Erros de Entrada",
            "Quantidade e percentual de registros com entrada inválida.",
            "Erros de entrada ÷ total × 100.",
        ),
        _dictionary_row(
            "Indicador",
            "Regra mais acionada",
            "Regra principal encontrada no maior número de registros.",
            "Counter.most_common(); em empate, prevalece a primeira ocorrência.",
        ),
        _dictionary_row(
            "Taxa",
            "Taxa de qualidade da entrada",
            "Percentual de registros sem erro de entrada.",
            "(Total − erros de entrada) ÷ total × 100; meta > 80%.",
        ),
        _dictionary_row(
            "Taxa",
            "Taxa de revisão humana",
            "Percentual de registros ambíguos.",
            "Ambíguos ÷ total × 100; meta < 15%.",
        ),
        _dictionary_row(
            "Taxa",
            "Taxa de retrabalho",
            "Percentual de registros divergentes.",
            "Divergências ÷ total × 100; meta < 6%.",
        ),
        _dictionary_row(
            "Indicador",
            "Ganho estimado de tempo",
            "Tempo poupado pela automação em comparação ao processo manual.",
            "Total × (2,0 min manual − 0,25 min automático).",
        ),
        _dictionary_row(
            "Sinalização",
            "✓",
            "O indicador atende à meta de referência.",
            "Comparação estrita com a meta exibida no cartão.",
        ),
        _dictionary_row(
            "Auditoria de ML",
            "Decisão de ML",
            "Predição complementar aplicada a um lote ambíguo.",
            "Uma linha por consulta ou fallback; usa a mesma fonte do log estruturado.",
        ),
        _dictionary_row(
            "Auditoria de ML",
            "REVISAO_ML_OFFLINE",
            "Fallback seguro quando a API ML está indisponível.",
            "Classe, probabilidade, confiança, ação e latência ficam vazias.",
        ),
        _dictionary_row(
            "Sinalização",
            "⚠",
            "O indicador requer atenção porque não atende à meta.",
            "Comparação estrita com a meta exibida no cartão.",
        ),
    ]

    for number in range(1, 13):
        code = f"RN{number:02d}"
        rows.append(
            _dictionary_row(
                "Regra de negócio",
                code,
                MOTIVOS.get(code, RESERVED_RULE_DESCRIPTION),
                "Aplicada pelo motor RN01–RN12.",
            )
        )
    return rows


def _dictionary_row(
    category: str,
    term: str,
    definition: str,
    formula: str,
) -> dict[str, str]:
    return dict(zip(DICTIONARY_COLUMNS, (category, term, definition, formula)))


def _frame_from_rows(rows: list[dict[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=list(BUSINESS_COLUMNS))


def _business_row(record: RegistroValidado) -> dict[str, Any]:
    original = record.campos_originais
    return {
        "Data de referência": _reference_date(record, original),
        "Lote": _value(original, "lote_id"),
        "Produto": _value(original, "produto"),
        "Linha": _value(original, "linha"),
        "Turno": _value(original, "turno"),
        "Status": record.status_normalizado,
        "Responsável": _value(original, "responsavel"),
        "Observação": _value(original, "observacao"),
        "Classificação": record.classificacao,
        "Motivo": record.motivo,
    }


def _value(original: Mapping[str, Any], key: str) -> Any:
    value = original.get(key, "")
    return "" if _is_missing(value) else value


def _reference_date(
    record: RegistroValidado,
    original: Mapping[str, Any],
) -> date | str:
    value = original.get("data_referencia") or record.data_referencia
    if _is_missing(value):
        return ""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    try:
        return date.fromisoformat(text)
    except ValueError:
        pass

    try:
        day, month, year = (int(part) for part in text.split("/"))
        return date(year, month, day)
    except ValueError:
        return text


def record_order_key(record: RegistroValidado) -> tuple[date, str, int]:
    reference = _reference_date(record, record.campos_originais)
    sortable_date = reference if isinstance(reference, date) else date.max
    line_order = record.linha_origem or _integer_value(
        record.campos_originais.get("ordem_linha")
    )
    return sortable_date, record.aba_origem, line_order


def _integer_value(value: Any) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _is_missing(value: Any) -> bool:
    if value is None:
        return True
    try:
        missing = pd.isna(value)
    except (TypeError, ValueError):
        return False
    return bool(missing) if isinstance(missing, bool) else False


def _validate_classifications(records: list[RegistroValidado]) -> None:
    invalid = sorted(
        {
            record.classificacao
            for record in records
            if record.classificacao not in CLASSIFICATION_SHEETS
        }
    )
    if invalid:
        values = ", ".join(repr(value) for value in invalid)
        raise ValueError(f"Classificações não suportadas pelo relatório: {values}")


def _format_summary_sheet(
    sheet: Any,
    records: list[RegistroValidado],
    indicators: OperationalIndicators,
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:J2")
    title = sheet["A1"]
    title.value = "Dashboard Executivo · Conferência de Lotes"
    title.fill = SUMMARY_TITLE_FILL
    title.font = Font(color="FFFFFF", bold=True, size=18)
    title.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24

    totals = Counter(record.classificacao for record in records)
    _write_summary_cards(sheet, indicators)
    _write_executive_cards(sheet, indicators)
    _write_classification_table(sheet, totals)
    daily_rows = _write_daily_table(sheet, records)
    _add_doughnut_chart(sheet)
    _add_line_chart(sheet, daily_rows)
    _format_summary_layout(sheet)


def _write_summary_cards(
    sheet: Any,
    indicators: OperationalIndicators,
) -> None:
    for (
        label_range,
        value_range,
        label,
        attribute,
        color,
        is_percentage,
    ) in SUMMARY_CARDS:
        sheet.merge_cells(label_range)
        sheet.merge_cells(value_range)
        label_cell = sheet[label_range.split(":")[0]]
        value_cell = sheet[value_range.split(":")[0]]
        label_cell.value = label
        label_cell.fill = PatternFill(fill_type="solid", fgColor=color)
        label_cell.font = Font(color="FFFFFF", bold=True, size=11)
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        value = getattr(indicators, attribute)
        value_cell.value = value / 100.0 if is_percentage else value
        value_cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        value_cell.font = Font(color="1F1F1F", bold=True, size=20)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_percentage:
            value_cell.number_format = "0.0%"


def _write_executive_cards(
    sheet: Any,
    indicators: OperationalIndicators,
) -> None:
    values = (
        _most_triggered_rule_text(indicators),
        _target_text(indicators.taxa_qualidade_entrada, 80.0, higher_is_better=True),
        _target_text(indicators.taxa_revisao_humana, 15.0, higher_is_better=False),
        _target_text(indicators.taxa_retrabalho, 6.0, higher_is_better=False),
        (
            f"{indicators.ganho_estimado_tempo_minutos:.2f} min | "
            f"{indicators.ganho_estimado_tempo_horas:.2f} h"
        ),
    )
    for (label_range, value_range, label, color), value in zip(
        EXECUTIVE_CARDS,
        values,
        strict=True,
    ):
        sheet.merge_cells(label_range)
        sheet.merge_cells(value_range)
        label_cell = sheet[label_range.split(":")[0]]
        value_cell = sheet[value_range.split(":")[0]]

        label_cell.value = label
        label_cell.fill = PatternFill(fill_type="solid", fgColor=color)
        label_cell.font = Font(color="FFFFFF", bold=True, size=10)
        label_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        value_cell.value = value
        value_cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        value_cell.font = Font(color="1F1F1F", bold=True, size=14)
        value_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _most_triggered_rule_text(indicators: OperationalIndicators) -> str:
    if indicators.regra_mais_acionada_codigo == "N/A":
        return "Nenhuma regra acionada"
    return (
        f"{indicators.regra_mais_acionada_codigo} · "
        f"{indicators.regra_mais_acionada_qtd} ocorrência(s)\n"
        f"{indicators.regra_mais_acionada_nome}"
    )


def _target_text(value: float, target: float, *, higher_is_better: bool) -> str:
    target_met = value > target if higher_is_better else value < target
    symbol = "✓" if target_met else "⚠"
    return f"{value:.1f}% {symbol}"


def _write_classification_table(sheet: Any, totals: Counter[str]) -> None:
    sheet["R1"] = "Classificação"
    sheet["S1"] = "Total"
    for row_index, classification in enumerate(CLASSIFICATION_SHEETS, start=2):
        sheet.cell(row=row_index, column=18, value=classification)
        sheet.cell(row=row_index, column=19, value=totals[classification])


def _write_daily_table(sheet: Any, records: list[RegistroValidado]) -> int:
    daily_counts: defaultdict[date, Counter[str]] = defaultdict(Counter)
    for record in records:
        reference = _reference_date(record, record.campos_originais)
        if isinstance(reference, date):
            daily_counts[reference][record.classificacao] += 1

    headers = (
        "Data",
        "Divergências",
        "Ambíguos",
        "Erros de Entrada",
        "Total de problemas",
    )
    for column_index, header in enumerate(headers, start=21):
        sheet.cell(row=1, column=column_index, value=header)

    dashboard_dates = sorted(daily_counts)[-MAX_DASHBOARD_DAYS:]
    for row_index, reference in enumerate(dashboard_dates, start=2):
        counts = daily_counts[reference]
        divergence = counts[CLASSIFICACAO_DIVERGENCIA]
        ambiguous = counts[CLASSIFICACAO_AMBIGUO]
        input_errors = counts[CLASSIFICACAO_ERRO_ENTRADA]
        values = (
            reference,
            divergence,
            ambiguous,
            input_errors,
            divergence + ambiguous + input_errors,
        )
        for column_index, value in enumerate(values, start=21):
            sheet.cell(row=row_index, column=column_index, value=value)
        sheet.cell(row=row_index, column=21).number_format = "dd/mm/yyyy"

    return len(dashboard_dates)


def _add_doughnut_chart(sheet: Any) -> None:
    chart = DoughnutChart()
    chart.title = "Distribuição por classificação"
    chart.style = 10
    chart.holeSize = 55
    chart.firstSliceAng = 270
    chart.legend.position = "r"
    chart.height = 8.2
    chart.width = 12.5

    data = Reference(sheet, min_col=19, min_row=1, max_row=5)
    labels = Reference(sheet, min_col=18, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.series[0].data_points = [
        DataPoint(
            idx=index,
            spPr=GraphicalProperties(solidFill=CLASSIFICATION_COLORS[classification]),
        )
        for index, classification in enumerate(CLASSIFICATION_SHEETS)
    ]
    sheet.add_chart(chart, "A25")


def _add_line_chart(sheet: Any, daily_rows: int) -> None:
    chart = LineChart()
    chart.title = "Evolução diária dos problemas"
    chart.style = 13
    chart.y_axis.title = "Registros"
    chart.x_axis.title = "Data"
    chart.x_axis.number_format = "dd/mm"
    chart.legend.position = "b"
    chart.height = 8.2
    chart.width = 15.5

    if daily_rows:
        data = Reference(
            sheet, min_col=22, max_col=25, min_row=1, max_row=daily_rows + 1
        )
        dates = Reference(sheet, min_col=21, min_row=2, max_row=daily_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)
        colors = ("C00000", "ED7D31", "7F8C8D", "4472C4")
        for series, color in zip(chart.series, colors, strict=True):
            series.graphicalProperties.line.solidFill = color
            series.graphicalProperties.line.width = 24000
            series.marker.symbol = "circle"
            series.marker.size = 6

    sheet.add_chart(chart, "G25")


def _format_summary_layout(sheet: Any) -> None:
    for column in range(1, 17):
        sheet.column_dimensions[get_column_letter(column)].width = 12

    for row in (4, 8, 12, 16, 20):
        sheet.row_dimensions[row].height = 24
    for row in (5, 6, 9, 10, 13, 14, 17, 18, 21, 22):
        sheet.row_dimensions[row].height = 22

    sheet.freeze_panes = "A3"
    sheet.print_area = "A1:P42"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_view.zoomScale = 85


def _format_data_sheet(sheet: Any) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 30

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet["A"][1:]:
        if isinstance(cell.value, (date, datetime)):
            cell.number_format = "dd/mm/yyyy"
            cell.alignment = Alignment(horizontal="center", vertical="top")

    for column_index, column_name in enumerate(BUSINESS_COLUMNS, start=1):
        values = [column_name]
        values.extend(
            "" if cell.value is None else str(cell.value)
            for cell in list(sheet.columns)[column_index - 1][1:]
        )
        content_width = max(len(value) for value in values) + 2
        maximum = 60 if column_name == "Motivo" else 40
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(content_width, 12), maximum
        )


def _format_table_sheet(
    sheet: Any,
    *,
    table_name: str,
    percentage_columns: tuple[int, ...] = (),
) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 30

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index in percentage_columns:
        for cell in list(sheet.columns)[column_index - 1][1:]:
            cell.number_format = "0.0%"

    for column_index, cells in enumerate(sheet.columns, start=1):
        content_width = max(
            len("" if cell.value is None else str(cell.value)) for cell in cells
        )
        maximum = 75 if column_index in {2, 3, 4} else 28
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(content_width + 2, 14), maximum
        )

    if sheet.max_row > 1:
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
