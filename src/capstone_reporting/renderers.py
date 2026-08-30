"""Renderizadores dos artefatos derivados do snapshot híbrido."""

from __future__ import annotations

from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from src.excel_reporting.models import RegistroValidado
from src.excel_reporting.report_writer import write_excel_report
from src.operational_indicators import calcular_indicadores

from .models import HybridReportItem, HybridReportSnapshot, describe_fallback

CAPSTONE_SHEET_NAME = "Pipeline Híbrido"
CAPSTONE_COLUMNS = (
    "Lote",
    "Classificação",
    "Status operacional",
    "Origem dos dados",
    "Coleta desktop",
    "Coleta web",
    "Origem da decisão",
    "Confiança ML",
    "Motivo do fallback",
    "Modo degradado",
    "ID da execução",
    "ID de correlação",
    "ID da task",
)


def write_capstone_excel(snapshot: HybridReportSnapshot, destination: Path) -> Path:
    """Preserva as nove abas e acrescenta a visão comum do pipeline híbrido."""
    records = [
        _validated_record(item, index)
        for index, item in enumerate(snapshot.items, start=1)
    ]
    indicators = calcular_indicadores(records)
    write_excel_report(
        records,
        indicators,
        destination,
        ml_decisions=snapshot.ml_decisions,
    )

    workbook = load_workbook(destination)
    if CAPSTONE_SHEET_NAME in workbook.sheetnames:
        del workbook[CAPSTONE_SHEET_NAME]
    sheet = workbook.create_sheet(CAPSTONE_SHEET_NAME)
    metadata = (
        ("Tipo do relatório", snapshot.report_type),
        ("Status", snapshot.status),
        ("ID da execução", snapshot.execution_id),
        ("ID de correlação", snapshot.correlation_id),
        ("ID da task raiz", snapshot.root_task_id),
        ("ID da task", snapshot.task_id),
        ("Total de itens", snapshot.total_items),
        ("Itens processados", snapshot.processed_items),
        ("Itens com falha", snapshot.failed_items),
        ("Itens para revisão", snapshot.review_items),
        ("Modo degradado", "Sim" if snapshot.modo_degradado else "Não"),
        ("Status ML", snapshot.ml_status),
    )
    for row in metadata:
        sheet.append(row)
    sheet.append(())
    sheet.append(CAPSTONE_COLUMNS)
    header_row = sheet.max_row
    for item in snapshot.items:
        sheet.append(_capstone_row(item))

    fill = PatternFill(fill_type="solid", fgColor="17365D")
    for cell in sheet[header_row]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(1, len(metadata) + 1):
        sheet.cell(row=row, column=1).font = Font(bold=True)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:M{sheet.max_row}"
    widths = (20, 20, 22, 24, 18, 18, 22, 16, 24, 17, 24, 24, 24)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column)].width = width
    workbook.save(destination)
    return destination


def write_capstone_markdown(snapshot: HybridReportSnapshot, destination: Path) -> Path:
    counts = snapshot.classification_counts
    lines = [
        "# Relatório do pipeline híbrido",
        "",
        "## Execução",
        "",
        "| Campo | Valor |",
        "|---|---|",
        f"| Tipo | {_md(snapshot.report_type)} |",
        f"| Status | {_md(snapshot.status)} |",
        f"| Execução | {_md(snapshot.execution_id)} |",
        f"| Correlação | {_md(snapshot.correlation_id)} |",
        f"| Task raiz | {_md(snapshot.root_task_id)} |",
        f"| Task | {_md(snapshot.task_id)} |",
        f"| Modo degradado | {'Sim' if snapshot.modo_degradado else 'Não'} |",
        f"| Motivo do fallback | {_md(_fallback_display(snapshot.motivo_fallback))} |",
        f"| Coleta desktop | {_md(_desktop_status(snapshot))} |",
        f"| Coleta web | {_md(_web_status(snapshot))} |",
        f"| ML | {_md(snapshot.ml_status)} |",
        "",
        "## Totais",
        "",
        "| Indicador | Quantidade |",
        "|---|---:|",
        f"| Total | {snapshot.total_items} |",
        f"| Processados | {snapshot.processed_items} |",
        f"| Falhas | {snapshot.failed_items} |",
        f"| Revisão humana | {snapshot.review_items} |",
    ]
    lines.extend(
        f"| {_md(classification)} | {quantity} |"
        for classification, quantity in counts.items()
    )
    lines.extend(
        [
            "",
            "## Decisões por item",
            "",
            "| Lote | Classificação | Status | Origem dos dados | Origem da decisão | Confiança ML | Fallback | Degradado |",
            "|---|---|---|---|---|---:|---|---|",
        ]
    )
    lines.extend(
        "| "
        + " | ".join(
            (
                _md(item.lote_id),
                _md(item.classificacao),
                _md(item.status_operacional),
                _md(", ".join(item.origem_dados)),
                _md(item.origem_decisao),
                "" if item.confianca_ml is None else f"{item.confianca_ml:.4f}",
                _md(_fallback_display(item.motivo_fallback)),
                "Sim" if item.modo_degradado else "Não",
            )
        )
        + " |"
        for item in snapshot.items
    )
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return destination


def write_capstone_pdf(snapshot: HybridReportSnapshot, destination: Path) -> Path:
    """Gera um PDF textual sem observações livres nem segredos."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        str(destination),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="Relatório do pipeline híbrido",
        author="relatorio-alertas-v2",
    )
    story: list[Any] = [
        Paragraph("Relatório do pipeline híbrido", styles["Title"]),
        Spacer(1, 4 * mm),
    ]
    summary_rows = [
        ["Campo", "Valor"],
        ["Tipo", snapshot.report_type],
        ["Status", snapshot.status],
        ["Execução", snapshot.execution_id],
        ["Correlação", snapshot.correlation_id],
        ["Task raiz", snapshot.root_task_id],
        ["Task", snapshot.task_id],
        ["Desktop", _desktop_status(snapshot)],
        ["Web", _web_status(snapshot)],
        ["ML", snapshot.ml_status],
        ["Modo degradado", "Sim" if snapshot.modo_degradado else "Não"],
        ["Fallback", _fallback_display(snapshot.motivo_fallback)],
        [
            "Totais",
            (
                f"{snapshot.total_items} total; {snapshot.processed_items} processados; "
                f"{snapshot.failed_items} falhas; {snapshot.review_items} revisões"
            ),
        ],
    ]
    summary_rows.extend(
        [classification, quantity]
        for classification, quantity in snapshot.classification_counts.items()
    )
    story.append(_pdf_table(summary_rows, (50 * mm, 215 * mm)))
    story.extend(
        [Spacer(1, 5 * mm), Paragraph("Decisões por item", styles["Heading2"])]
    )
    item_rows: list[list[Any]] = [[
        "Lote",
        "Classificação",
        "Status",
        "Origem dos dados",
        "Origem da decisão",
        "Confiança ML",
        "Fallback",
        "Degradado",
    ]]
    item_rows.extend(
        [
            item.lote_id,
            item.classificacao,
            item.status_operacional,
            ", ".join(item.origem_dados),
            item.origem_decisao,
            "" if item.confianca_ml is None else f"{item.confianca_ml:.4f}",
            _fallback_display(item.motivo_fallback),
            "Sim" if item.modo_degradado else "Não",
        ]
        for item in snapshot.items
    )
    story.append(
        _pdf_table(
            item_rows,
            (
                28 * mm,
                32 * mm,
                30 * mm,
                35 * mm,
                35 * mm,
                24 * mm,
                45 * mm,
                20 * mm,
            ),
        )
    )
    document.build(story)
    return destination


def _validated_record(item: HybridReportItem, index: int) -> RegistroValidado:
    fields = dict(item.campos_relatorio)
    fields["lote_id"] = item.lote_id
    fields.pop("observacao", None)
    return RegistroValidado(
        campos_originais=fields,
        status_original=item.status_operacional,
        status_normalizado=item.status_operacional,
        classificacao=item.classificacao,
        motivo=item.motivo_fallback or item.classificacao,
        regras_violadas=item.regras_violadas,
        data_referencia=str(fields.get("data_referencia") or ""),
        aba_origem="Pipeline Capstone",
        linha_origem=index,
        regra_aplicada=item.regra_aplicada,
    )


def _capstone_row(item: HybridReportItem) -> tuple[Any, ...]:
    return (
        item.lote_id,
        item.classificacao,
        item.status_operacional,
        ", ".join(item.origem_dados),
        item.status_coleta_desktop,
        item.status_coleta_web,
        item.origem_decisao,
        item.confianca_ml,
        _fallback_display(item.motivo_fallback),
        "Sim" if item.modo_degradado else "Não",
        item.execution_id,
        item.correlation_id,
        item.task_id,
    )


def _pdf_table(rows: list[list[Any]], widths: tuple[float, ...]) -> Table:
    rendered = [
        [Paragraph(escape(str(value or "")), getSampleStyleSheet()["BodyText"]) for value in row]
        for row in rows
    ]
    table = Table(rendered, colWidths=widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
            ]
        )
    )
    return table


def _desktop_status(snapshot: HybridReportSnapshot) -> str:
    return _status_by_alias(snapshot, "desktop", "estoque", "estoque-desktop-v1")


def _web_status(snapshot: HybridReportSnapshot) -> str:
    return _status_by_alias(snapshot, "web", "pedidos", "fornecedores-web-v1")


def _status_by_alias(snapshot: HybridReportSnapshot, *aliases: str) -> str:
    statuses = {key.casefold(): value for key, value in snapshot.source_statuses.items()}
    for alias in aliases:
        if alias.casefold() in statuses:
            return statuses[alias.casefold()]
    return "UNAVAILABLE"


def _md(value: object) -> str:
    return str(value).replace("|", "\\|").replace("\r", " ").replace("\n", " ")


def _fallback_display(value: str | None) -> str:
    if value is None:
        return describe_fallback(None)
    return f"{value}: {describe_fallback(value)}"


__all__ = [
    "CAPSTONE_SHEET_NAME",
    "write_capstone_excel",
    "write_capstone_markdown",
    "write_capstone_pdf",
]
