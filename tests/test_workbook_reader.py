from collections import Counter
from pathlib import Path

from openpyxl import Workbook

from src.excel_reporting.workbook_reader import (
    DEFAULT_WORKBOOK_PATH,
    list_daily_sheet_names,
    read_workbook,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
REAL_WORKBOOK_PATH = PROJECT_ROOT / "dados_entrada" / "inspecao_lotes_10dias.xlsx"

DAILY_HEADER = [
    "lote_id",
    "produto",
    "linha",
    "turno",
    "status",
    "responsavel",
    "data",
    "observacao",
]


def _create_workbook(path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    _add_daily_sheet(
        workbook,
        "Insp_15_06_2026",
        [
            ["L001", "Monitor", "L1", "A", "APROVADO", "Ana", "15/06/2026", ""],
            ["L002", "TV", "L2", "B", "APROVADO", "Bia", "15/06/2026", ""],
            ["", "TV", "L2", "B", "APROVADO", "Bia", "15/06/2026", ""],
            ["", "Ar", "L3", "C", "APROVADO", "Dora", "15/06/2026", ""],
            ["L001", "Monitor", "L1", "A", "APROVADO", "Caio", "15/06/2026", ""],
        ],
    )
    _add_daily_sheet(
        workbook,
        "Insp_16_06_2026",
        [
            ["L001", "Monitor", "L1", "A", "APROVADO", "Ana", "16/06/2026", ""],
            ["L003", "Ar", "L3", "C", "REPROVADO", "Bia", "16/06/2026", "Falha"],
            ["L003", "Ar", "L3", "C", "REPROVADO", "Caio", "16/06/2026", "Falha"],
        ],
    )
    _add_daily_sheet(
        workbook,
        "Resumo",
        [["decorativo", "", "", "", "", "", "", ""]],
    )
    _add_reference_sheet(workbook)
    workbook.save(path)


def _add_daily_sheet(workbook, title, rows):
    sheet = workbook.create_sheet(title)
    sheet.append([f"PLANILHA DE INSPECAO - {title}"])
    sheet.append(["Arquivo gerado automaticamente"])
    sheet.append(DAILY_HEADER)
    for row in rows:
        sheet.append(row)
    sheet.append([])
    sheet.append([f"Total de registros: {len(rows)}"])


def _add_reference_sheet(workbook):
    sheet = workbook.create_sheet("Base_Referencia")
    sheet.append(["BASE DE REFERENCIA DE LOTES"])
    sheet.append(["lote_id", "codigo_produto", "descricao_produto", "status_cadastro"])
    sheet.append(["L001", "MON", "Monitor", "Ativo"])
    sheet.append(["L002", "TV", "Televisao", "Ativo"])
    sheet.append(["L003", "AR", "Ar-condicionado", "Ativo"])
    sheet.append(["Nota final: registro decorativo"])


def test_list_daily_sheet_names_uses_expected_pattern(tmp_path):
    workbook_path = tmp_path / "inspecao_lotes.xlsx"
    _create_workbook(workbook_path)

    assert list_daily_sheet_names(workbook_path) == [
        "Insp_15_06_2026",
        "Insp_16_06_2026",
    ]


def test_read_workbook_consolidates_daily_sheets_and_reference_base(tmp_path):
    workbook_path = tmp_path / "inspecao_lotes.xlsx"
    _create_workbook(workbook_path)

    result = read_workbook(workbook_path)

    assert len(result.registros) == 8
    assert {row["lote_id"] for row in result.base_referencia} == {"L001", "L002", "L003"}
    assert result.lotes_referencia == {"L001", "L002", "L003"}
    assert any(row["lote_id"] == "" for row in result.registros)
    assert all(not row["lote_id"].startswith("Total de registros") for row in result.registros)
    assert all(not row["lote_id"].startswith("Nota final") for row in result.base_referencia)


def test_read_workbook_adds_origin_date_and_original_order(tmp_path):
    workbook_path = tmp_path / "inspecao_lotes.xlsx"
    _create_workbook(workbook_path)

    result = read_workbook(workbook_path)

    first = result.registros[0]
    assert first["aba_origem"] == "Insp_15_06_2026"
    assert first["data_referencia"] == "2026-06-15"
    assert first["ordem_linha"] == 1
    assert result.registros[5]["aba_origem"] == "Insp_16_06_2026"
    assert result.registros[5]["data_referencia"] == "2026-06-16"
    assert result.registros[5]["ordem_linha"] == 1


def test_daily_duplicate_counter_restarts_for_each_sheet(tmp_path):
    workbook_path = tmp_path / "inspecao_lotes.xlsx"
    _create_workbook(workbook_path)

    result = read_workbook(workbook_path)
    duplicates = [
        (row["aba_origem"], row["lote_id"], row["ordem_linha"])
        for row in result.registros
        if row["duplicado_no_dia"]
    ]

    assert duplicates == [
        ("Insp_15_06_2026", "L001", 5),
        ("Insp_16_06_2026", "L003", 3),
    ]
    assert result.total_duplicidades_adicionais == 2
    assert result.contagem_lotes_por_aba == {
        "Insp_15_06_2026": Counter({"L001": 2, "L002": 1}),
        "Insp_16_06_2026": Counter({"L003": 2, "L001": 1}),
    }
    assert not result.registros[2]["duplicado_no_dia"]
    assert not result.registros[3]["duplicado_no_dia"]
    assert result.registros[2]["ocorrencia_lote_no_dia"] == 0
    assert result.registros[3]["ocorrencia_lote_no_dia"] == 0
    assert not result.registros[5]["duplicado_no_dia"]


def test_real_workbook_matches_acceptance_numbers():
    result = read_workbook(REAL_WORKBOOK_PATH)
    registros_por_aba = Counter(row["aba_origem"] for row in result.registros)

    assert len(list_daily_sheet_names(REAL_WORKBOOK_PATH)) == 10
    assert set(registros_por_aba.values()) == {25}
    assert len(result.registros) == 250
    assert result.total_duplicidades_adicionais == 20


def test_default_workbook_path_is_relative():
    assert not DEFAULT_WORKBOOK_PATH.is_absolute()
