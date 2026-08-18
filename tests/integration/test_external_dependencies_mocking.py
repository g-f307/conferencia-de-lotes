"""Integra leitura, validacao e relatorio com dependencias controladas."""

from __future__ import annotations

import socket
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook

from src.excel_reporting import REPORT_SHEET_NAMES, gerar_relatorio_excel
from src.excel_reporting import service as service_module
from src.excel_reporting import workbook_reader as workbook_reader_module
from src.vault_client import BotCityVaultProvider

pytestmark = pytest.mark.integration


class FixedDateTime(datetime):
    """Relogio deterministico usado somente nos testes do log."""

    @classmethod
    def now(cls, tz=None):
        fixed = cls(2026, 8, 16, 12, 30, 45)
        return fixed if tz is None else fixed.replace(tzinfo=tz)


def test_fluxo_controlado_integra_leitura_validacao_relatorio_e_log(
    workbook_sintetico: Path,
    base_referencia_simulada: list[dict[str, object]],
    diretorio_saida: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Arrange
    output_path = diretorio_saida / "relatorio.xlsx"
    log_path = diretorio_saida / "execucao.log"
    base_mock = MagicMock(return_value=base_referencia_simulada)
    perf_counter_mock = MagicMock(side_effect=(100.0, 101.25))
    write_report_mock = MagicMock(wraps=service_module.write_excel_report)
    markdown_mock = MagicMock(wraps=service_module.gerar_resumo_executivo)
    calcular_mock = MagicMock(wraps=service_module.calcular_indicadores)
    network_mock = MagicMock(side_effect=AssertionError("rede nao permitida"))
    credential_mock = MagicMock(
        side_effect=AssertionError("credencial real nao permitida")
    )
    monkeypatch.setattr(
        workbook_reader_module,
        "read_reference_base",
        base_mock,
    )
    monkeypatch.setattr(service_module.time, "perf_counter", perf_counter_mock)
    monkeypatch.setattr(service_module, "datetime", FixedDateTime)
    monkeypatch.setattr(service_module, "write_excel_report", write_report_mock)
    monkeypatch.setattr(service_module, "gerar_resumo_executivo", markdown_mock)
    monkeypatch.setattr(service_module, "calcular_indicadores", calcular_mock)
    monkeypatch.setattr(socket, "create_connection", network_mock)
    monkeypatch.setattr(
        BotCityVaultProvider,
        "get_credential",
        credential_mock,
    )

    # Act
    result = gerar_relatorio_excel(
        workbook_sintetico,
        output_path,
        log_path=log_path,
    )

    # Assert
    workbook = load_workbook(output_path)
    log_text = log_path.read_text(encoding="utf-8")
    temporary_output = Path(write_report_mock.call_args.args[2])

    assert workbook.sheetnames == list(REPORT_SHEET_NAMES)
    assert result.total_registros == 6
    assert result.total_classificacoes == 6
    assert result.validos == 2
    assert result.divergencias == 2
    assert result.ambiguos == 1
    assert result.erros_entrada == 1
    assert result.regras == {"RN05": 1, "RN09": 1, "RN10": 1, "RN12": 1}
    assert len(result.registros_validados) == 6
    assert result.duracao_segundos == 1.25
    assert output_path.is_file()
    assert log_path.is_file()
    assert output_path.parent == diretorio_saida
    assert log_path.parent == diretorio_saida
    assert temporary_output.parent == diretorio_saida
    assert not temporary_output.exists()
    assert "data_hora=2026-08-16T12:30:45" in log_text
    assert "duracao_segundos=1.250" in log_text
    assert "total_registros=6" in log_text
    assert "taxa_qualidade_entrada=" in log_text
    assert "ganho_estimado_tempo_minutos=" in log_text
    assert "regra_mais_acionada_descricao=" in log_text
    assert "regra_mais_acionada_qtd=" in log_text
    assert "ganho_estimado_tempo_horas=" in log_text
    assert (diretorio_saida / "resumo_executivo.md").is_file()
    base_mock.assert_called_once_with(workbook_sintetico)
    assert perf_counter_mock.call_count == 2
    write_report_mock.assert_called_once()
    markdown_mock.assert_called_once()
    calcular_mock.assert_called_once()

    excel_indicators = write_report_mock.call_args.args[1]
    markdown_indicators = markdown_mock.call_args.args[0]
    assert excel_indicators is markdown_indicators

    network_mock.assert_not_called()
    credential_mock.assert_not_called()
    workbook.close()


def test_falha_na_escrita_remove_arquivo_temporario(
    workbook_sintetico: Path,
    base_referencia_simulada: list[dict[str, object]],
    diretorio_saida: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Arrange
    output_path = diretorio_saida / "relatorio.xlsx"
    log_path = diretorio_saida / "execucao.log"
    base_mock = MagicMock(return_value=base_referencia_simulada)

    def interromper_escrita(_registros, _indicators, temporary_path):
        Path(temporary_path).write_bytes(b"arquivo parcial")
        raise OSError("falha controlada de escrita")

    write_report_mock = MagicMock(side_effect=interromper_escrita)
    monkeypatch.setattr(
        workbook_reader_module,
        "read_reference_base",
        base_mock,
    )
    monkeypatch.setattr(service_module.time, "perf_counter", MagicMock(return_value=1.0))
    monkeypatch.setattr(service_module, "write_excel_report", write_report_mock)

    # Act / Assert
    with pytest.raises(OSError, match="falha controlada de escrita"):
        gerar_relatorio_excel(
            workbook_sintetico,
            output_path,
            log_path=log_path,
        )

    assert not output_path.exists()
    assert not log_path.exists()
    assert list(diretorio_saida.glob("*.tmp.xlsx")) == []
    base_mock.assert_called_once_with(workbook_sintetico)
    write_report_mock.assert_called_once()


def test_falha_na_geracao_markdown_previne_publicacao_do_excel(
    workbook_sintetico: Path,
    base_referencia_simulada: list[dict[str, object]],
    diretorio_saida: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    output_path = diretorio_saida / "relatorio.xlsx"
    log_path = diretorio_saida / "execucao.log"
    markdown_path = diretorio_saida / "resumo_executivo.md"
    base_mock = MagicMock(return_value=base_referencia_simulada)

    def interromper_markdown(_indicadores, temporary_path):
        Path(temporary_path).write_text("markdown parcial")
        raise OSError("falha ao gerar markdown")

    markdown_mock = MagicMock(side_effect=interromper_markdown)
    monkeypatch.setattr(workbook_reader_module, "read_reference_base", base_mock)
    monkeypatch.setattr(service_module, "gerar_resumo_executivo", markdown_mock)

    with pytest.raises(OSError, match="falha ao gerar markdown"):
        gerar_relatorio_excel(
            workbook_sintetico,
            output_path,
            log_path=log_path,
        )

    assert not output_path.exists()
    assert not markdown_path.exists()
    assert not list(diretorio_saida.glob("*.tmp"))
