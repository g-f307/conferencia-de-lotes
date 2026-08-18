from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.chart import DoughnutChart, LineChart
from openpyxl.workbook.workbook import Workbook

from src.excel_reporting import (
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO_ENTRADA,
    CLASSIFICACAO_VALIDO,
    DICTIONARY_COLUMNS,
    RANKING_COLUMNS,
    REPORT_SHEET_NAMES,
    RegistroValidado,
    write_excel_report,
)
from src.excel_reporting.report_writer import record_order_key
from src.markdown_reporting import gerar_resumo_executivo
from src.operational_indicators import OperationalIndicators, calcular_indicadores

pytestmark = pytest.mark.integration


@dataclass(frozen=True)
class ConsolidatedArtifacts:
    workbook: Workbook
    excel_path: Path
    markdown_path: Path
    indicators: OperationalIndicators


def _record(
    sequence: int,
    classification: str,
    *,
    rule: str = "",
) -> RegistroValidado:
    reference = date(2026, 6, 15)
    return RegistroValidado(
        campos_originais={
            "lote_id": f"L{sequence:03d}",
            "produto": "Monitor",
            "linha": "L1",
            "turno": "A",
            "status": "APROVADO",
            "responsavel": "Ana",
            "observacao": "",
            "data_referencia": reference,
        },
        status_original="APROVADO",
        status_normalizado="APROVADO",
        classificacao=classification,
        motivo=f"{rule}: cenário controlado" if rule else "Registro válido",
        regras_violadas=(rule,) if rule else (),
        data_referencia=reference,
        aba_origem="Insp_15_06_2026",
        linha_origem=sequence,
        regra_aplicada=rule,
    )


@pytest.fixture
def consolidated_artifacts(tmp_path):
    records = [
        _record(1, CLASSIFICACAO_DIVERGENCIA, rule="RN05"),
        _record(2, CLASSIFICACAO_DIVERGENCIA, rule="RN05"),
        _record(3, CLASSIFICACAO_AMBIGUO, rule="RN09"),
        _record(4, CLASSIFICACAO_ERRO_ENTRADA, rule="RN01"),
        *[_record(sequence, CLASSIFICACAO_VALIDO) for sequence in range(5, 10)],
    ]
    ordered = sorted(records, key=record_order_key)
    indicators = calcular_indicadores(ordered)
    excel_path = tmp_path / "relatorio_conferencia_lotes.xlsx"
    markdown_path = tmp_path / "resumo_executivo.md"
    write_excel_report(ordered, indicators, excel_path)
    gerar_resumo_executivo(indicators, markdown_path)

    workbook = load_workbook(excel_path)
    yield ConsolidatedArtifacts(
        workbook=workbook,
        excel_path=excel_path,
        markdown_path=markdown_path,
        indicators=indicators,
    )
    workbook.close()


@pytest.fixture
def consolidated_workbook(consolidated_artifacts):
    return consolidated_artifacts.workbook


def test_gera_artefatos_fisicos_e_exatamente_oito_abas(consolidated_artifacts):
    assert consolidated_artifacts.excel_path.is_file()
    assert consolidated_artifacts.excel_path.stat().st_size > 0
    assert consolidated_artifacts.markdown_path.is_file()
    assert consolidated_artifacts.markdown_path.stat().st_size > 0
    assert consolidated_artifacts.workbook.sheetnames == list(REPORT_SHEET_NAMES)


def test_abas_operacionais_nao_misturam_classificacoes(consolidated_workbook):
    expected = {
        "Válidos": CLASSIFICACAO_VALIDO,
        "Divergências": CLASSIFICACAO_DIVERGENCIA,
        "Ambíguos": CLASSIFICACAO_AMBIGUO,
        "Erros de Entrada": CLASSIFICACAO_ERRO_ENTRADA,
    }
    for sheet_name, classification in expected.items():
        values = {cell.value for cell in consolidated_workbook[sheet_name]["I"][1:]}
        assert values == {classification}


def test_dashboard_exibe_indicadores_metas_e_graficos_nativos(
    consolidated_workbook,
):
    summary = consolidated_workbook["Resumo"]

    assert summary["A5"].value == 9
    assert summary["E5"].value == 5
    assert summary["I5"].value == pytest.approx(5 / 9, abs=0.0001)
    assert summary["A9"].value == 2
    assert summary["E9"].value == pytest.approx(2 / 9, abs=0.0001)
    assert summary["I9"].value == 1
    assert summary["A13"].value == pytest.approx(1 / 9, abs=0.0001)
    assert summary["E13"].value == 1
    assert summary["I13"].value == pytest.approx(1 / 9, abs=0.0001)
    assert summary["M5"].value.startswith("RN05 · 2 ocorrência(s)")
    assert summary["M9"].value == "88.9% ✓"
    assert summary["M13"].value == "11.1% ✓"
    assert summary["M17"].value == "22.2% ⚠"
    assert summary["M21"].value == "15.75 min | 0.26 h"

    assert len(summary._charts) == 2
    assert any(isinstance(chart, DoughnutChart) for chart in summary._charts)
    assert any(isinstance(chart, LineChart) for chart in summary._charts)
    assert summary._images == []


def test_ranking_usa_a_mesma_regra_principal_do_dashboard(consolidated_workbook):
    summary = consolidated_workbook["Resumo"]
    ranking = consolidated_workbook["Ranking de Regras"]

    assert tuple(cell.value for cell in ranking[1]) == RANKING_COLUMNS
    assert [cell.value for cell in ranking["A"][1:]] == ["RN05", "RN09", "RN01"]
    assert [cell.value for cell in ranking["C"][1:]] == [2, 1, 1]
    assert ranking["D2"].value == pytest.approx(2 / 9, abs=0.0001)
    assert summary["M5"].value.startswith(
        f"{ranking['A2'].value} · {ranking['C2'].value} ocorrência(s)"
    )


def test_dicionario_cobre_termos_formulas_e_todas_as_regras(consolidated_workbook):
    dictionary = consolidated_workbook["Dicionário"]
    rows = list(dictionary.iter_rows(min_row=2, values_only=True))
    terms = {row[1] for row in rows}

    assert tuple(cell.value for cell in dictionary[1]) == DICTIONARY_COLUMNS
    assert {
        CLASSIFICACAO_VALIDO,
        CLASSIFICACAO_DIVERGENCIA,
        CLASSIFICACAO_AMBIGUO,
        CLASSIFICACAO_ERRO_ENTRADA,
        "Taxa de qualidade da entrada",
        "Taxa de revisão humana",
        "Taxa de retrabalho",
        "Ganho estimado de tempo",
    } <= terms
    assert {f"RN{number:02d}" for number in range(1, 13)} <= terms
    assert all(row[2] and row[3] for row in rows)


def test_resumo_executivo_fisico_contem_os_dez_indicadores(
    consolidated_artifacts,
):
    conteudo = consolidated_artifacts.markdown_path.read_text(encoding="utf-8")

    assert "# Resumo Executivo: Conferência de Lotes" in conteudo
    assert "| Total de Registros Processados | 9 |" in conteudo
    assert "| Cadastros Válidos | 5 (55.6%) |" in conteudo
    assert "| Divergências Identificadas | 2 (22.2%) |" in conteudo
    assert "| Casos Ambíguos (Revisão Manual) | 1 (11.1%) |" in conteudo
    assert "| Erros de Entrada | 1 (11.1%) |" in conteudo
    assert "| Regra Mais Acionada | RN05 (2 ocorrências) |" in conteudo
    assert "| Taxa de Qualidade da Entrada | 88.9% |" in conteudo
    assert "| Taxa de Revisão Humana | 11.1% |" in conteudo
    assert "| Taxa de Retrabalho | 22.2% |" in conteudo
    assert "| Ganho Estimado de Tempo | 15.75 min &#124; 0.26 h |" in conteudo
    assert "2,0 minutos" in conteudo
    assert "0,25 minutos" in conteudo
    assert "Observação Metodológica:" in conteudo


def test_regra_aplicada_alimenta_indicador_seis_e_ranking(tmp_path):
    records = [
        _record(1, CLASSIFICACAO_DIVERGENCIA, rule="RN05"),
        _record(2, CLASSIFICACAO_DIVERGENCIA, rule="RN05"),
        _record(3, CLASSIFICACAO_AMBIGUO, rule="RN09"),
    ]
    records_without_main_rule = [
        replace(record, regra_aplicada="") for record in records
    ]
    ordered = sorted(records_without_main_rule, key=record_order_key)
    indicators = calcular_indicadores(ordered)
    output = tmp_path / "sem-regra-aplicada.xlsx"

    write_excel_report(ordered, indicators, output)
    workbook = load_workbook(output)
    try:
        assert indicators.regra_mais_acionada_codigo == "N/A"
        assert workbook["Resumo"]["M5"].value == "Nenhuma regra acionada"
        assert workbook["Ranking de Regras"].max_row == 1
    finally:
        workbook.close()
