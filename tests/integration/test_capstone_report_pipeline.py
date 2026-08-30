from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.capstone_reporting import CapstoneReportService
from src.capstone_reporting.main import CapstoneReportSettings, run
from src.capstone_reporting.renderers import CAPSTONE_SHEET_NAME
from src.excel_reporting.report_writer import REPORT_SHEET_NAMES

pytestmark = pytest.mark.integration


def _validation(
    lote_id: str,
    classification: str,
    status: str,
    rule: str = "",
) -> dict[str, object]:
    return {
        "campos_originais": {
            "data_referencia": "2026-08-30",
            "lote_id": lote_id,
            "produto": "Produto A",
            "linha": "Linha 1",
            "turno": "Manhã",
            "responsavel": "Operação",
            "observacao": "não transportar esta observação",
        },
        "status_original": status,
        "status_normalizado": status,
        "classificacao": classification,
        "motivo": rule or "Registro válido",
        "regras_violadas": [rule] if rule else [],
        "data_referencia": "2026-08-30",
        "aba_origem": "Dia 1",
        "linha_origem": 2,
        "regra_aplicada": rule,
    }


def _decision() -> dict[str, object]:
    return {
        "timestamp": "2026-08-30T12:00:00+00:00",
        "execution_id": "exec-report-001",
        "bot_id": "classificador-ml-v1",
        "lote_id": "L002",
        "classe": "divergencia_quantidade",
        "probabilidade": 0.88,
        "nivel_confianca": "alta",
        "acao": "revisar",
        "resultado_aplicado": "DIVERGENCIA",
        "latencia_ms": 18.0,
        "causa_provavel": "divergencia_quantidade",
        "origem_decisao": "ml",
        "confianca_ml": 0.88,
        "motivo_fallback": None,
    }


def _payload() -> dict[str, object]:
    decision = _decision()
    records = [
        {
            "lote_id": "L001",
            "status_operacional": "APROVADO",
            "classificacao": "Válido",
            "regras_violadas": [],
            "regra_aplicada": "",
            "origens_consultadas": ["estoque", "pedidos", "validacao"],
            "fontes_ausentes": [],
            "modo_degradado": False,
            "validacao": _validation("L001", "Válido", "APROVADO"),
        },
        {
            "lote_id": "L002",
            "status_operacional": "DIVERGENCIA",
            "classificacao": "Divergência",
            "regras_violadas": ["RN05"],
            "regra_aplicada": "RN05",
            "origens_consultadas": ["estoque", "pedidos", "validacao"],
            "fontes_ausentes": [],
            "modo_degradado": False,
            "validacao": _validation("L002", "Divergência", "REPROVADO", "RN05"),
        },
    ]
    return {
        "report_type": "BUSINESS",
        "task_id": "task-report-001",
        "source_statuses": {
            "estoque-desktop-v1": "AVAILABLE",
            "fornecedores-web-v1": "AVAILABLE",
        },
        "consolidation_result": {
            "status": "SUCCESS",
            "modo_degradado": False,
            "payload": {
                "records": records,
                "item_failures": [],
                "total_items": 2,
                "processed_items": 2,
                "failed_items": 0,
                "review_items": 0,
            },
        },
        "ml_result": {
            "status": "SUCCESS",
            "execution_id": "exec-report-001",
            "correlation_id": "corr-report-001",
            "root_task_id": "task-root-001",
            "task_id": "task-ml-001",
            "modo_degradado": False,
            "payload": {
                "records": [
                    {"lote_id": "L001", "decisao_ml": None},
                    {"lote_id": "L002", "decisao_ml": decision},
                ],
                "ml_decisions": [decision],
            },
        },
    }


def test_relatorio_capstone_gera_artefatos_coerentes_sem_misturar_abas(
    tmp_path: Path,
) -> None:
    result = CapstoneReportService(tmp_path).generate(_payload())

    assert result.paths.excel is not None
    assert result.paths.excel.is_file()
    assert result.paths.markdown.is_file()
    assert result.paths.pdf.stat().st_size > 0
    assert result.paths.summary.is_file()

    summary = json.loads(result.paths.summary.read_text(encoding="utf-8"))
    assert summary["total_items"] == 2
    assert summary["processed_items"] == 2
    assert summary["classification_counts"] == {
        "Válido": 1,
        "Divergência": 1,
        "Ambíguo": 0,
        "Erro de Entrada": 0,
    }
    assert summary["items"][1]["origem_decisao"] == "ml"
    assert summary["items"][1]["confianca_ml"] == pytest.approx(0.88)

    markdown = result.paths.markdown.read_text(encoding="utf-8")
    assert "| Total | 2 |" in markdown
    assert "| Processados | 2 |" in markdown
    assert "L002 | Divergência | DIVERGENCIA | estoque, pedidos, validacao | ml | 0.8800" in markdown
    assert "não transportar esta observação" not in markdown

    workbook = load_workbook(result.paths.excel, data_only=False)
    assert workbook.sheetnames == [*REPORT_SHEET_NAMES, CAPSTONE_SHEET_NAME]
    assert workbook["Válidos"].max_row == 2
    assert workbook["Válidos"]["I2"].value == "Válido"
    assert workbook["Divergências"].max_row == 2
    assert workbook["Divergências"]["I2"].value == "Divergência"
    assert workbook["Ambíguos"].max_row == 1
    assert workbook["Erros de Entrada"].max_row == 1
    assert workbook[CAPSTONE_SHEET_NAME]["A7"].value == "Total de itens"
    assert workbook[CAPSTONE_SHEET_NAME]["B7"].value == 2
    assert "não transportar esta observação" not in result.paths.excel.read_bytes().decode(
        "latin-1",
        errors="ignore",
    )


def test_entrypoint_independente_processa_envelope_configurado(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    input_path = tmp_path / "pipeline-capstone.json"
    input_path.write_text(
        json.dumps(_payload(), ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("ALERTS_ENABLED", "false")
    monkeypatch.setenv("MAESTRO_ENABLED", "false")
    monkeypatch.setenv("ORCHESTRATION_ENABLED", "false")
    monkeypatch.setenv("WEB_AUTOMATION_ENABLED", "false")
    monkeypatch.setenv("ML_ENABLED", "false")
    monkeypatch.setenv("LOG_FILE", str(tmp_path / "execucao.log"))

    result = run(
        CapstoneReportSettings(
            input_path=input_path,
            output_dir=tmp_path / "relatorios",
            degraded_alert_seconds=300,
        )
    )

    assert result["status"] == "SUCCESS"
    assert result["execution_id"] == "exec-report-001"
    assert Path(result["summary_path"]).is_file()
