from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.excel_reporting import (
    BUSINESS_COLUMNS,
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO_ENTRADA,
    CLASSIFICACAO_VALIDO,
    CLASSIFICATION_SHEETS,
    DATA_SHEET_NAMES,
    ML_DECISION_COLUMNS,
    ML_DECISIONS_SHEET_NAME,
    REPORT_SHEET_NAMES,
    RegistroValidado,
    ValidationService,
    read_workbook,
    write_excel_report,
)
from src.ml_audit import MLDecisionAudit

PROJECT_ROOT = Path(__file__).resolve().parents[2]
REAL_WORKBOOK_PATH = PROJECT_ROOT / "dados_entrada" / "inspecao_lotes_10dias.xlsx"
pytestmark = pytest.mark.integration


def _record(
    *,
    lote: str,
    classification: str,
    status_original: str = "APROVADO",
    status_normalizado: str = "APROVADO",
    reason: str = "Registro válido pelas regras RN01-RN12",
    rules: tuple[str, ...] = (),
    reference_date: str = "2026-06-15",
    sheet: str = "Insp_15_06_2026",
    line: int = 1,
) -> RegistroValidado:
    return RegistroValidado(
        campos_originais={
            "lote_id": lote,
            "produto": "Monitor",
            "linha": "L1",
            "turno": "A",
            "status": status_original,
            "responsavel": "Ana",
            "data": "15/06/2026",
            "observacao": "",
            "data_referencia": reference_date,
            "aba_origem": sheet,
            "ordem_linha": line,
            "duplicado_no_dia": False,
            "ocorrencia_lote_no_dia": 1,
        },
        status_original=status_original,
        status_normalizado=status_normalizado,
        classificacao=classification,
        motivo=reason,
        regras_violadas=rules,
        data_referencia="15/06/2026",
        aba_origem=sheet,
        linha_origem=line,
        regra_aplicada=rules[0] if rules else "",
    )


@pytest.fixture
def classified_records() -> list[RegistroValidado]:
    return [
        _record(lote="L006", classification=CLASSIFICACAO_VALIDO, line=6),
        _record(
            lote="L003",
            classification=CLASSIFICACAO_DIVERGENCIA,
            reason="RN05: Lote não encontrado na base de referência",
            rules=("RN05",),
            line=3,
        ),
        _record(
            lote="L005",
            classification=CLASSIFICACAO_ERRO_ENTRADA,
            reason="RN12: Data ausente ou fora do formato DD/MM/AAAA",
            rules=("RN12",),
            line=5,
        ),
        _record(
            lote="L004",
            classification=CLASSIFICACAO_AMBIGUO,
            status_original="EM AJUSTE",
            status_normalizado="EM AJUSTE",
            reason="RN09: Status desconhecido e não normalizável",
            rules=("RN09",),
            line=4,
        ),
        _record(
            lote="L002",
            classification=CLASSIFICACAO_DIVERGENCIA,
            status_original="NOK",
            status_normalizado="REPROVADO",
            reason="RN10: Lote reprovado sem observação",
            rules=("RN10",),
            line=2,
        ),
        _record(
            lote="L001",
            classification=CLASSIFICACAO_VALIDO,
            status_original="OK",
            status_normalizado="APROVADO",
            reference_date="2026-06-14",
            sheet="Insp_14_06_2026",
            line=1,
        ),
    ]


def _load_generated_report(tmp_path, records):
    from src.excel_reporting.report_writer import record_order_key
    from src.operational_indicators import calcular_indicadores
    output = tmp_path / "relatorio_conferencia_lotes.xlsx"
    ordered = sorted(records, key=record_order_key)
    indicators = calcular_indicadores(ordered)
    returned_path = write_excel_report(ordered, indicators, output)
    assert returned_path == output
    return output, load_workbook(output)


def test_report_contains_exactly_nine_required_sheets(tmp_path, classified_records):
    _, workbook = _load_generated_report(tmp_path, classified_records)

    assert workbook.sheetnames == list(REPORT_SHEET_NAMES)
    assert workbook.sheetnames[-1] == ML_DECISIONS_SHEET_NAME


def test_ml_sheet_is_created_with_headers_when_there_are_no_decisions(
    tmp_path,
    classified_records,
):
    _, workbook = _load_generated_report(tmp_path, classified_records)
    sheet = workbook[ML_DECISIONS_SHEET_NAME]

    assert tuple(cell.value for cell in sheet[1]) == ML_DECISION_COLUMNS
    assert sheet.max_row == 1
    assert sheet.freeze_panes == "A2"


def test_ml_sheet_uses_audit_records_and_preserves_numeric_fields(
    tmp_path,
    classified_records,
):
    from src.excel_reporting.report_writer import record_order_key
    from src.operational_indicators import calcular_indicadores

    output = tmp_path / "relatorio_ml.xlsx"
    ordered = sorted(classified_records, key=record_order_key)
    indicators = calcular_indicadores(ordered)
    decisions = [
        MLDecisionAudit(
            timestamp="2026-08-19T12:30:00+00:00",
            execution_id="exec-123",
            bot_id="bot-ml",
            lote_id="L001",
            classe="valido_automatico",
            probabilidade=0.91,
            nivel_confianca="alta",
            acao="valido_automatico",
            resultado_aplicado="APROVADO",
            latencia_ms=22.75,
        ),
        MLDecisionAudit(
            timestamp="2026-08-19T12:31:00+00:00",
            execution_id="exec-123",
            bot_id="bot-ml",
            lote_id="L002",
            classe=None,
            probabilidade=None,
            nivel_confianca=None,
            acao=None,
            resultado_aplicado="REVISAO_ML_OFFLINE",
            latencia_ms=None,
        ),
    ]

    write_excel_report(ordered, indicators, output, ml_decisions=decisions)
    workbook = load_workbook(output)
    sheet = workbook[ML_DECISIONS_SHEET_NAME]

    assert sheet.max_row == 3
    assert sheet["D2"].value == "L001"
    assert sheet["F2"].value == pytest.approx(0.91)
    assert sheet["J2"].value == pytest.approx(22.75)
    assert sheet["I3"].value == "REVISAO_ML_OFFLINE"
    assert all(sheet.cell(3, column).value is None for column in (5, 6, 7, 8, 10))


def test_all_and_classification_sheets_have_expected_records(
    tmp_path,
    classified_records,
):
    _, workbook = _load_generated_report(tmp_path, classified_records)
    expected_counts = {
        "Todos": 6,
        "Válidos": 2,
        "Divergências": 2,
        "Ambíguos": 1,
        "Erros de Entrada": 1,
    }

    for sheet_name, expected_count in expected_counts.items():
        assert workbook[sheet_name].max_row - 1 == expected_count

    classified_total = sum(
        workbook[sheet_name].max_row - 1
        for sheet_name in CLASSIFICATION_SHEETS.values()
    )
    assert classified_total == workbook["Todos"].max_row - 1


def test_each_sheet_contains_only_its_classification(tmp_path, classified_records):
    _, workbook = _load_generated_report(tmp_path, classified_records)
    expected = {
        "Válidos": CLASSIFICACAO_VALIDO,
        "Divergências": CLASSIFICACAO_DIVERGENCIA,
        "Ambíguos": CLASSIFICACAO_AMBIGUO,
        "Erros de Entrada": CLASSIFICACAO_ERRO_ENTRADA,
    }
    classification_column = BUSINESS_COLUMNS.index("Classificação") + 1

    for sheet_name, classification in expected.items():
        values = {
            cell.value
            for cell in list(workbook[sheet_name].columns)[classification_column - 1][1:]
        }
        assert values == {classification}


def test_columns_are_business_friendly_and_hide_technical_metadata(
    tmp_path,
    classified_records,
):
    _, workbook = _load_generated_report(tmp_path, classified_records)

    for sheet_name in DATA_SHEET_NAMES:
        headers = tuple(cell.value for cell in workbook[sheet_name][1])
        assert headers == BUSINESS_COLUMNS
        assert "regras_violadas" not in headers
        assert "aba_origem" not in headers
        assert "ordem_linha" not in headers


def test_final_status_is_normalized_and_records_keep_chronological_order(
    tmp_path,
    classified_records,
):
    _, workbook = _load_generated_report(tmp_path, classified_records)
    sheet = workbook["Todos"]
    lote_column = BUSINESS_COLUMNS.index("Lote") + 1
    status_column = BUSINESS_COLUMNS.index("Status") + 1
    date_column = BUSINESS_COLUMNS.index("Data de referência") + 1

    lotes = [row[lote_column - 1].value for row in sheet.iter_rows(min_row=2)]
    statuses = [row[status_column - 1].value for row in sheet.iter_rows(min_row=2)]
    dates = [row[date_column - 1].value for row in sheet.iter_rows(min_row=2)]

    assert lotes == ["L001", "L002", "L003", "L004", "L005", "L006"]
    assert "OK" not in statuses
    assert "NOK" not in statuses
    assert statuses[:2] == ["APROVADO", "REPROVADO"]
    assert all(isinstance(value, datetime) for value in dates)
    assert dates == sorted(dates)


def test_data_sheets_receive_filters_freeze_panes_and_visual_formatting(
    tmp_path,
    classified_records,
):
    _, workbook = _load_generated_report(tmp_path, classified_records)

    for sheet_name in DATA_SHEET_NAMES:
        sheet = workbook[sheet_name]
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref == sheet.dimensions
        assert sheet["A1"].fill.fgColor.rgb == "001F4E78"
        assert sheet["A1"].font.bold
        assert sheet["A1"].font.color.rgb == "00FFFFFF"
        assert sheet.column_dimensions["J"].width >= 12
        assert sheet["A2"].number_format == "dd/mm/yyyy"


def test_summary_is_prepared_without_extra_technical_sheets(
    tmp_path,
    classified_records,
):
    _, workbook = _load_generated_report(tmp_path, classified_records)
    summary = workbook["Resumo"]

    assert summary["A1"].value == "Dashboard Executivo · Conferência de Lotes"
    assert "A1:J2" in {str(cell_range) for cell_range in summary.merged_cells.ranges}
    assert not summary.sheet_view.showGridLines


def test_unknown_classification_fails_instead_of_losing_a_record(tmp_path):
    invalid_record = _record(lote="L999", classification="Nova classificação")
    from src.operational_indicators import calcular_indicadores
    indicators = calcular_indicadores([invalid_record])

    with pytest.raises(ValueError, match="Nova classificação"):
        write_excel_report([invalid_record], indicators, tmp_path / "relatorio.xlsx")


def test_real_workbook_generates_250_rows_without_mixing_categories(tmp_path):
    source = read_workbook(REAL_WORKBOOK_PATH)
    service = ValidationService(source.lotes_referencia)
    validated = [
        service.validar_registro(
            record,
            aba_origem=str(record["aba_origem"]),
            linha_origem=int(record["ordem_linha"]),
        )
        for record in source.registros
    ]

    output, workbook = _load_generated_report(tmp_path, validated)

    assert output.is_file()
    assert workbook["Todos"].max_row - 1 == 250
    assert sum(
        workbook[sheet_name].max_row - 1
        for sheet_name in CLASSIFICATION_SHEETS.values()
    ) == 250
    assert {record.classificacao for record in validated} == {
        CLASSIFICACAO_VALIDO,
        CLASSIFICACAO_DIVERGENCIA,
        CLASSIFICACAO_AMBIGUO,
        CLASSIFICACAO_ERRO_ENTRADA,
    }
