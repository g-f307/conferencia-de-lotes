"""E2E controlado do pipeline Excel, sem navegador ou servicos externos."""

from __future__ import annotations

import socket
from collections import Counter
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import DoughnutChart, LineChart

from src.excel_reporting import (
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO_ENTRADA,
    CLASSIFICACAO_VALIDO,
    CLASSIFICATION_SHEETS,
    REPORT_SHEET_NAMES,
    gerar_relatorio_excel,
)
from src.excel_reporting import service as service_module
from src.vault_client import BotCityVaultProvider

pytestmark = pytest.mark.e2e

DAILY_HEADERS = (
    "lote_id",
    "produto",
    "linha",
    "turno",
    "status",
    "responsavel",
    "data",
    "observacao",
)


class FixedDateTime(datetime):
    """Relogio deterministico usado na evidencia do log E2E."""

    @classmethod
    def now(cls, tz=None):
        fixed = cls(2026, 8, 17, 9, 30, 15)
        return fixed if tz is None else fixed.replace(tzinfo=tz)


def _registro(
    lote_id: str,
    *,
    data: str,
    produto: str = "Produto controlado",
    status: str = "APROVADO",
    observacao: str = "",
) -> dict[str, str]:
    return {
        "lote_id": lote_id,
        "produto": produto,
        "linha": "Linha A",
        "turno": "Manha",
        "status": status,
        "responsavel": "Equipe E2E",
        "data": data,
        "observacao": observacao,
    }


def _append_daily_sheet(
    workbook: Workbook,
    sheet_name: str,
    records: list[dict[str, str]],
) -> None:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(["Inspecao controlada de lotes"])
    sheet.append([])
    sheet.append(DAILY_HEADERS)
    for record in records:
        sheet.append([record[header] for header in DAILY_HEADERS])


def _create_controlled_workbook(path: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _append_daily_sheet(
        workbook,
        "Insp_10_08_2026",
        [
            _registro("L001", data="10/08/2026"),
            _registro("L002", data="10/08/2026", status="REPROVADO"),
            _registro("L003", data="10/08/2026", status="em analise"),
            _registro("L004", data="10/08/2026", produto=""),
            _registro("L005", data="10/08/2026"),
            _registro("L005", data="10/08/2026"),
        ],
    )
    _append_daily_sheet(
        workbook,
        "Insp_11_08_2026",
        [_registro("L005", data="11/08/2026")],
    )

    reference = workbook.create_sheet("Base_Referencia")
    reference.append(["Base de Referencia controlada"])
    reference.append(["lote_id", "produto"])
    for lote_id in ("L001", "L002", "L003", "L004", "L005"):
        reference.append([lote_id, "Produto controlado"])

    workbook.save(path)
    workbook.close()
    return path


def _sheet_rows(sheet) -> list[tuple[object, ...]]:
    return list(sheet.iter_rows(min_row=2, max_col=10, values_only=True))


def test_pipeline_excel_completo_com_dados_controlados(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    # Arrange
    input_path = _create_controlled_workbook(tmp_path / "entrada_controlada.xlsx")
    output_path = tmp_path / "relatorio_controlado.xlsx"
    log_path = tmp_path / "execucao_controlada.log"
    perf_counter_mock = MagicMock(side_effect=(50.0, 51.5))
    network_mock = MagicMock(side_effect=AssertionError("rede nao permitida no E2E"))
    credential_mock = MagicMock(
        side_effect=AssertionError("credencial real nao permitida no E2E")
    )
    monkeypatch.setattr(service_module.time, "perf_counter", perf_counter_mock)
    monkeypatch.setattr(service_module, "datetime", FixedDateTime)
    monkeypatch.setattr(socket, "create_connection", network_mock)
    monkeypatch.setattr(BotCityVaultProvider, "get_credential", credential_mock)

    # Act
    result = gerar_relatorio_excel(input_path, output_path, log_path=log_path)

    # Assert
    workbook = load_workbook(output_path)
    summary = workbook["Resumo"]
    all_rows = _sheet_rows(workbook["Todos"])
    classification_counts = Counter(row[8] for row in all_rows)
    classified_rows = Counter(
        row
        for sheet_name in CLASSIFICATION_SHEETS.values()
        for row in _sheet_rows(workbook[sheet_name])
    )
    percentages = (
        summary["I5"].value,
        summary["E9"].value,
        summary["A13"].value,
        summary["I13"].value,
    )

    assert workbook.sheetnames == list(REPORT_SHEET_NAMES)
    assert result.total_registros == result.total_classificacoes == 7
    assert classification_counts == {
        CLASSIFICACAO_VALIDO: 3,
        CLASSIFICACAO_DIVERGENCIA: 2,
        CLASSIFICACAO_AMBIGUO: 1,
        CLASSIFICACAO_ERRO_ENTRADA: 1,
    }
    assert summary["A5"].value == 7
    assert summary["E5"].value == 3
    assert summary["A9"].value == 2
    assert summary["I9"].value == 1
    assert summary["E13"].value == 1
    assert sum(percentages) == pytest.approx(1.0, abs=0.0002)
    assert Counter(all_rows) == classified_rows
    assert len(summary._charts) == 2
    assert sum(isinstance(chart, DoughnutChart) for chart in summary._charts) == 1
    assert sum(isinstance(chart, LineChart) for chart in summary._charts) == 1

    l005_records = [
        record
        for record in result.registros_validados
        if record["campos_originais"]["lote_id"] == "L005"
    ]
    assert [record["aba_origem"] for record in l005_records] == [
        "Insp_10_08_2026",
        "Insp_10_08_2026",
        "Insp_11_08_2026",
    ]
    assert ["RN11" in record["regras_violadas"] for record in l005_records] == [
        False,
        True,
        False,
    ]

    assert input_path.parent == output_path.parent == log_path.parent == tmp_path
    assert output_path.is_file()
    assert log_path.is_file()
    assert list(tmp_path.glob("*.tmp.xlsx")) == []
    assert "data_hora=2026-08-17T09:30:15" in log_path.read_text(encoding="utf-8")
    assert "duracao_segundos=1.500" in log_path.read_text(encoding="utf-8")
    assert perf_counter_mock.call_count == 2
    network_mock.assert_not_called()
    credential_mock.assert_not_called()
    workbook.close()
