from __future__ import annotations

from collections import Counter
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.chart import DoughnutChart, LineChart

from gerar_relatorio import main
from src.excel_reporting import REPORT_SHEET_NAMES
from src.excel_reporting.service import gerar_relatorio_excel


PROJECT_ROOT = Path(__file__).resolve().parents[2]
REAL_WORKBOOK_PATH = PROJECT_ROOT / "dados_entrada" / "inspecao_lotes_10dias.xlsx"
pytestmark = pytest.mark.e2e


def test_report_service_processes_real_dataset_end_to_end(tmp_path):
    output = tmp_path / "relatorio_conferencia_lotes.xlsx"
    log_path = tmp_path / "execucao_relatorio.log"

    result = gerar_relatorio_excel(REAL_WORKBOOK_PATH, output, log_path=log_path)
    workbook = load_workbook(output)
    summary = workbook["Resumo"]
    classifications = Counter(
        row[8]
        for row in workbook["Todos"].iter_rows(min_row=2, values_only=True)
    )

    assert output.is_file()
    assert log_path.is_file()
    assert result.total_registros == 250
    assert result.total_classificacoes == result.total_registros
    assert workbook.sheetnames == list(REPORT_SHEET_NAMES)
    assert len(summary._charts) == 2
    assert any(isinstance(chart, DoughnutChart) for chart in summary._charts)
    assert any(isinstance(chart, LineChart) for chart in summary._charts)
    assert sum(classifications.values()) == 250
    assert "Total de registros: 25" not in {
        row[1] for row in workbook["Todos"].iter_rows(min_row=2, values_only=True)
    }
    assert result.regras["RN11"] == 20


def test_report_service_writes_execution_log_with_same_totals(tmp_path):
    output = tmp_path / "relatorio.xlsx"
    log_path = tmp_path / "execucao_relatorio.log"

    result = gerar_relatorio_excel(REAL_WORKBOOK_PATH, output, log_path=log_path)
    log_text = log_path.read_text(encoding="utf-8")

    assert f"arquivo_processado={REAL_WORKBOOK_PATH}" in log_text
    assert f"total_registros={result.total_registros}" in log_text
    assert f"validos={result.validos}" in log_text
    assert f"divergencias={result.divergencias}" in log_text
    assert f"ambiguos={result.ambiguos}" in log_text
    assert f"erros_entrada={result.erros_entrada}" in log_text
    assert f"relatorio={output}" in log_text


def test_report_service_fails_fast_when_input_is_missing(tmp_path):
    output = tmp_path / "relatorio.xlsx"
    log_path = tmp_path / "execucao_relatorio.log"

    with pytest.raises(FileNotFoundError):
        gerar_relatorio_excel(tmp_path / "ausente.xlsx", output, log_path=log_path)

    assert not output.exists()
    assert not log_path.exists()
    assert list(tmp_path.glob("*.tmp.xlsx")) == []


def test_report_service_rejects_non_excel_input(tmp_path):
    input_path = tmp_path / "entrada.csv"
    input_path.write_text("lote_id\nL001\n", encoding="utf-8")

    with pytest.raises(ValueError, match="extensao Excel"):
        gerar_relatorio_excel(input_path, tmp_path / "relatorio.xlsx")


def test_cli_generates_report_and_prints_summary(tmp_path, capsys):
    output = tmp_path / "relatorio_cli.xlsx"
    log_path = tmp_path / "execucao_cli.log"

    exit_code = main(
        [
            "--entrada",
            str(REAL_WORKBOOK_PATH),
            "--saida",
            str(output),
            "--log",
            str(log_path),
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 0
    assert output.is_file()
    assert log_path.is_file()
    assert "Relatorio gerado com sucesso" in captured.out
    assert "Total de registros: 250" in captured.out


def test_cli_reads_ml_decisions_from_execution_summary(tmp_path, capsys):
    output = tmp_path / "relatorio_cli_ml.xlsx"
    log_path = tmp_path / "execucao_cli_ml.log"
    summary_path = tmp_path / "resumo_execucao.json"
    summary_path.write_text(
        json.dumps(
            {
                "ml_decisions": [
                    {
                        "timestamp": "2026-08-19T12:30:00+00:00",
                        "execution_id": "exec-cli",
                        "bot_id": "bot-ml",
                        "lote_id": "L001",
                        "classe": "revisar",
                        "probabilidade": 0.72,
                        "nivel_confianca": "media",
                        "acao": "revisar",
                        "resultado_aplicado": "REVISAO",
                        "latencia_ms": 18.4,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    exit_code = main(
        [
            "--entrada",
            str(REAL_WORKBOOK_PATH),
            "--saida",
            str(output),
            "--log",
            str(log_path),
            "--decisoes-ml",
            str(summary_path),
        ]
    )
    captured = capsys.readouterr()
    workbook = load_workbook(output)

    assert exit_code == 0
    assert workbook["Decisões de ML"].max_row == 2
    assert workbook["Decisões de ML"]["D2"].value == "L001"
    assert "Decisões de ML: 1" in captured.out
    workbook.close()


def test_cli_rejeita_decisao_ml_sem_rastreabilidade(tmp_path, capsys):
    output = tmp_path / "relatorio_cli_invalido.xlsx"
    summary_path = tmp_path / "resumo_invalido.json"
    summary_path.write_text(
        json.dumps(
            {
                "ml_decisions": [
                    {
                        "timestamp": "2026-08-19T12:30:00+00:00",
                        "bot_id": "bot-ml",
                        "lote_id": "L001",
                        "classe": "revisar",
                        "probabilidade": 0.72,
                        "nivel_confianca": "media",
                        "acao": "revisar",
                        "resultado_aplicado": "REVISAO",
                        "latencia_ms": 18.4,
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    exit_code = main(
        [
            "--entrada",
            str(REAL_WORKBOOK_PATH),
            "--saida",
            str(output),
            "--decisoes-ml",
            str(summary_path),
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 1
    assert "execution_id é obrigatório" in captured.err
    assert not output.exists()


def test_cli_returns_error_for_invalid_input(tmp_path, capsys):
    output = tmp_path / "relatorio_cli.xlsx"

    exit_code = main(
        [
            "--entrada",
            str(tmp_path / "nao_existe.xlsx"),
            "--saida",
            str(output),
        ]
    )
    captured = capsys.readouterr()

    assert exit_code == 1
    assert "Arquivo de entrada inexistente" in captured.err
    assert not output.exists()
