from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.chart import DoughnutChart, LineChart

from src.excel_reporting import (
    CLASSIFICACAO_AMBIGUO,
    CLASSIFICACAO_DIVERGENCIA,
    CLASSIFICACAO_ERRO_ENTRADA,
    CLASSIFICACAO_VALIDO,
    REPORT_SHEET_NAMES,
    RegistroValidado,
    ValidationService,
    read_workbook,
    write_excel_report,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
REAL_WORKBOOK_PATH = PROJECT_ROOT / "dados_entrada" / "inspecao_lotes_10dias.xlsx"


@pytest.fixture(scope="module")
def dashboard_workbook(tmp_path_factory):
    source = read_workbook(REAL_WORKBOOK_PATH)
    service = ValidationService(source.lotes_referencia)
    validated = [
        service.validar_registro(
            record,
            linha_origem=int(record["ordem_linha"]),
        )
        for record in source.registros
    ]
    output = tmp_path_factory.mktemp("dashboard") / "relatorio_conferencia_lotes.xlsx"
    write_excel_report(validated, output)
    return load_workbook(output)


def _summary_values(sheet):
    return {
        sheet[label_cell].value: sheet[value_cell].value
        for label_cell, value_cell in (
            ("A4", "A5"),
            ("E4", "E5"),
            ("I4", "I5"),
            ("A8", "A9"),
            ("E8", "E9"),
            ("I8", "I9"),
            ("A12", "A13"),
            ("E12", "E13"),
            ("I12", "I13"),
        )
    }


def test_dashboard_indicators_match_classification_sheets(dashboard_workbook):
    summary = dashboard_workbook["Resumo"]
    values = _summary_values(summary)
    counts = {
        CLASSIFICACAO_VALIDO: dashboard_workbook["Válidos"].max_row - 1,
        CLASSIFICACAO_DIVERGENCIA: dashboard_workbook["Divergências"].max_row - 1,
        CLASSIFICACAO_AMBIGUO: dashboard_workbook["Ambíguos"].max_row - 1,
        CLASSIFICACAO_ERRO_ENTRADA: dashboard_workbook["Erros de Entrada"].max_row - 1,
    }
    total = dashboard_workbook["Todos"].max_row - 1

    assert values["Total de registros"] == total == 250
    assert values["Total de válidos"] == counts[CLASSIFICACAO_VALIDO]
    assert values["Total de divergências"] == counts[CLASSIFICACAO_DIVERGENCIA]
    assert values["Total de ambíguos"] == counts[CLASSIFICACAO_AMBIGUO]
    assert values["Total de erros de entrada"] == counts[CLASSIFICACAO_ERRO_ENTRADA]
    assert values["% de válidos"] == pytest.approx(counts[CLASSIFICACAO_VALIDO] / total)
    assert values["% de divergências"] == pytest.approx(
        counts[CLASSIFICACAO_DIVERGENCIA] / total
    )
    assert values["% de ambíguos"] == pytest.approx(
        counts[CLASSIFICACAO_AMBIGUO] / total
    )
    assert values["% de erros de entrada"] == pytest.approx(
        counts[CLASSIFICACAO_ERRO_ENTRADA] / total
    )


def test_dashboard_percentages_total_one_hundred_percent(dashboard_workbook):
    values = _summary_values(dashboard_workbook["Resumo"])
    percentages = [value for label, value in values.items() if label.startswith("%")]

    assert sum(percentages) == pytest.approx(1.0)
    percentage_cells = ("I5", "E9", "A13", "I13")
    assert all(
        dashboard_workbook["Resumo"][cell].number_format == "0.0%"
        for cell in percentage_cells
    )


def test_dashboard_contains_native_doughnut_and_line_charts(dashboard_workbook):
    summary = dashboard_workbook["Resumo"]

    assert len(summary._charts) == 2
    doughnut_charts = [
        chart for chart in summary._charts if isinstance(chart, DoughnutChart)
    ]
    assert len(doughnut_charts) == 1
    assert (
        len([chart for chart in summary._charts if isinstance(chart, LineChart)]) == 1
    )
    assert summary._images == []


def test_doughnut_chart_references_four_classifications(dashboard_workbook):
    summary = dashboard_workbook["Resumo"]
    chart = next(chart for chart in summary._charts if isinstance(chart, DoughnutChart))

    assert [summary.cell(row=row, column=18).value for row in range(2, 6)] == [
        CLASSIFICACAO_VALIDO,
        CLASSIFICACAO_DIVERGENCIA,
        CLASSIFICACAO_AMBIGUO,
        CLASSIFICACAO_ERRO_ENTRADA,
    ]
    assert chart.series[0].val.numRef.f == "'Resumo'!$S$2:$S$5"
    assert chart.series[0].cat.numRef.f == "'Resumo'!$R$2:$R$5"
    assert len(chart.series[0].data_points) == 4


def test_line_chart_contains_ten_chronological_days_and_required_series(
    dashboard_workbook,
):
    summary = dashboard_workbook["Resumo"]
    chart = next(chart for chart in summary._charts if isinstance(chart, LineChart))
    dates = [summary.cell(row=row, column=21).value for row in range(2, 12)]
    headers = [summary.cell(row=1, column=column).value for column in range(22, 26)]

    assert len(dates) == 10
    assert all(isinstance(value, datetime) for value in dates)
    assert dates == sorted(dates)
    assert headers == [
        "Divergências",
        "Ambíguos",
        "Erros de Entrada",
        "Total de problemas",
    ]
    assert len(chart.series) == 4
    assert chart.series[0].val.numRef.f == "'Resumo'!$V$2:$V$11"
    assert chart.series[1].val.numRef.f == "'Resumo'!$W$2:$W$11"


def test_dashboard_limits_evolution_to_ten_most_recent_days(tmp_path):
    first_day = date(2026, 6, 1)
    dates = [first_day + timedelta(days=offset) for offset in range(11)]
    records = [
        RegistroValidado(
            campos_originais={"lote_id": f"LOTE-{index:02d}"},
            status_original="REPROVADO",
            status_normalizado="REPROVADO",
            classificacao=CLASSIFICACAO_DIVERGENCIA,
            motivo="RN05: Lote não encontrado na base de referência",
            regras_violadas=("RN05",),
            data_referencia=reference.isoformat(),
            aba_origem=f"Insp_{reference:%d_%m_%Y}",
            linha_origem=index,
        )
        for index, reference in enumerate(dates, start=1)
    ]
    output = tmp_path / "dashboard-11-dias.xlsx"

    write_excel_report(records, output)
    summary = load_workbook(output)["Resumo"]
    chart_dates = [
        summary.cell(row=row, column=21).value.date() for row in range(2, 12)
    ]

    assert chart_dates == dates[1:]
    assert summary["U12"].value is None


def test_daily_dashboard_values_match_all_records(dashboard_workbook):
    summary = dashboard_workbook["Resumo"]
    todos = dashboard_workbook["Todos"]
    classification_column = 9
    date_column = 1
    expected = {}
    for row in todos.iter_rows(min_row=2, values_only=True):
        day = row[date_column - 1]
        classification = row[classification_column - 1]
        expected.setdefault(day, {"Divergência": 0, "Ambíguo": 0, "Erro de Entrada": 0})
        if classification in expected[day]:
            expected[day][classification] += 1

    for row_index in range(2, 12):
        day = summary.cell(row=row_index, column=21).value
        divergence = summary.cell(row=row_index, column=22).value
        ambiguous = summary.cell(row=row_index, column=23).value
        input_errors = summary.cell(row=row_index, column=24).value
        total_problems = summary.cell(row=row_index, column=25).value
        counts = expected[day]

        assert divergence == counts["Divergência"]
        assert ambiguous == counts["Ambíguo"]
        assert input_errors == counts["Erro de Entrada"]
        assert total_problems == divergence + ambiguous + input_errors


def test_dashboard_keeps_six_sheets_and_auxiliary_tables_outside_print_area(
    dashboard_workbook,
):
    summary = dashboard_workbook["Resumo"]

    assert dashboard_workbook.sheetnames == list(REPORT_SHEET_NAMES)
    assert summary["R1"].value == "Classificação"
    assert summary["U1"].value == "Data"
    assert str(summary.print_area) == "'Resumo'!$A$1:$P$34"
    assert summary.page_setup.orientation == "landscape"
